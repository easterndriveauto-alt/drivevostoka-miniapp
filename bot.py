"""
bot.py — Telegram-интерфейс агента DriveVostoka
Обрабатывает сообщения и передаёт их в agent.py
"""

import asyncio
import logging
import os
import re

import anthropic
from dotenv import load_dotenv
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, InputMediaPhoto, Update, WebAppInfo
from telegram.constants import ParseMode
from telegram.ext import (
    Application, CallbackQueryHandler, CommandHandler,
    ContextTypes, MessageHandler, filters,
)

from io import BytesIO

from agent import run_agent
from image_utils import prepare_photos
from database import save_profile
from database import (
    clear_history, get_active_watchlist, get_all_leads, get_lead_by_id,
    get_leads, get_profile, get_stats, init_db, migrate_db,
    update_lead_status, upsert_client,
)
from monitor import run_monitor

load_dotenv()
logging.basicConfig(
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

TG_TOKEN        = os.environ["TG_BOT_TOKEN"]
CLAUDE_KEY      = os.environ["ANTHROPIC_API_KEY"]
MANAGER_CHAT_ID = os.environ.get("MANAGER_CHAT_ID", "")

claude = anthropic.AsyncAnthropic(api_key=CLAUDE_KEY)

BUDGET_MAP = {
    "seg_b_2m": (2_000_000,  "до 2 млн ₽"),
    "seg_b_3m": (3_000_000,  "2–3 млн ₽"),
    "seg_b_5m": (5_000_000,  "3–5 млн ₽"),
    "seg_b_5p": (8_000_000,  "5+ млн ₽"),
}

TYPE_MAP = {
    "seg_t_sedan": ("Седан",              "седан",    False),
    "seg_t_suv":   ("Кроссовер",          "кроссовер",False),
    "seg_t_ev":    ("Электро / Гибрид",   "электро",  True),
}

GOAL_MAP = {
    "seg_g_self":    "для себя",
    "seg_g_resale":  "на перепродажу",
    "seg_g_compare": "сравнить с рынком РФ",
}

MANAGER_URL  = "https://t.me/DriveVostoka"
MINI_APP_URL = os.environ.get("MINI_APP_URL", "")


async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    user    = update.effective_user

    upsert_client(
        chat_id=chat_id,
        username=user.username or "",
        first_name=user.first_name or "",
    )
    clear_history(chat_id)
    context.user_data.clear()

    miniapp_row = []
    if MINI_APP_URL:
        miniapp_row = [InlineKeyboardButton(
            "📱 Открыть каталог авто",
            web_app=WebAppInfo(url=MINI_APP_URL)
        )]

    rows = []
    if miniapp_row:
        rows.append(miniapp_row)
    rows += [
        [InlineKeyboardButton("🚗 Подобрать через бота",    callback_data="flow_start")],
        [InlineKeyboardButton("🚢 Как проходит доставка",   callback_data="flow_delivery")],
        [InlineKeyboardButton("💬 Написать менеджеру",      url=MANAGER_URL)],
    ]
    keyboard = InlineKeyboardMarkup(rows)

    await update.message.reply_text(
        f"👋 <b>Привет{', ' + user.first_name if user.first_name else ''}!</b>\n\n"
        "Я помогу привезти авто из Южной Кореи "
        "<b>с экономией до 30%</b> — даже после повышения утильсбора.\n\n"
        "Ответь на 3 вопроса — покажу реальные варианты под твой бюджет 👇",
        parse_mode=ParseMode.HTML,
        reply_markup=keyboard,
    )


async def _show_budget_step(message):
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("до 2 млн",  callback_data="seg_b_2m"),
         InlineKeyboardButton("2–3 млн",   callback_data="seg_b_3m")],
        [InlineKeyboardButton("3–5 млн",   callback_data="seg_b_5m"),
         InlineKeyboardButton("5+ млн",    callback_data="seg_b_5p")],
    ])
    await message.reply_text(
        "💰 <b>Вопрос 1 из 3.</b> Какой бюджет рассматриваешь?\n"
        "<i>(итоговая цена со всеми расходами до Владивостока)</i>",
        parse_mode=ParseMode.HTML,
        reply_markup=keyboard,
    )


async def _show_type_step(message):
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("🚗 Седан",            callback_data="seg_t_sedan")],
        [InlineKeyboardButton("🚙 Кроссовер / SUV",  callback_data="seg_t_suv")],
        [InlineKeyboardButton("⚡ Электро / Гибрид", callback_data="seg_t_ev")],
    ])
    await message.reply_text(
        "🚘 <b>Вопрос 2 из 3.</b> Какой тип авто интересует?",
        parse_mode=ParseMode.HTML,
        reply_markup=keyboard,
    )


async def _show_goal_step(message):
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("👤 Для себя",              callback_data="seg_g_self")],
        [InlineKeyboardButton("💼 На перепродажу",        callback_data="seg_g_resale")],
        [InlineKeyboardButton("📊 Сравнить с рынком РФ", callback_data="seg_g_compare")],
    ])
    await message.reply_text(
        "🎯 <b>Вопрос 3 из 3.</b> Какова цель покупки?",
        parse_mode=ParseMode.HTML,
        reply_markup=keyboard,
    )


async def _show_warmup(message, budget_label: str, type_label: str, goal: str):
    goal_insert = {
        "для себя":           "Берёшь для себя — значит важно сочетание цены, надёжности и комплектации.",
        "на перепродажу":     "Берёшь на перепродажу — значит важна ликвидность и маржа. Подберём варианты с хорошим спредом.",
        "сравнить с рынком РФ": "Хочешь сравнить — правильно делаешь. Цифры говорят сами за себя.",
    }.get(goal, "")

    await message.reply_text(
        "📌 <b>Важно знать перед поиском:</b>\n\n"
        "Из-за изменения утильсбора часть машин действительно подорожала.\n\n"
        "Но важно другое — <b>мы работаем с сегментами, которые почти не попали под рост.</b>\n\n"
        "✅ Прямые закупки с корейских аукционов\n"
        "✅ Без лишних посредников\n"
        "✅ Заранее считаем итоговую цену в РФ\n\n"
        f"{goal_insert}\n\n"
        "<b>В итоге: авто всё ещё выходит выгоднее рынка РФ.</b>\n\n"
        "Показываю реальные примеры 👇",
        parse_mode=ParseMode.HTML,
    )

    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("🔍 Получить подборку",    callback_data="flow_getlist")],
        [InlineKeyboardButton("💬 Написать менеджеру",   url=MANAGER_URL)],
    ])
    await message.reply_text(
        f"Хочу подобрать <b>3 варианта</b> под твой бюджет <b>{budget_label}</b> — {type_label}.\n\n"
        "Без оплаты. Просто покажу реальные цифры.",
        parse_mode=ParseMode.HTML,
        reply_markup=keyboard,
    )


async def _show_lead_block(message):
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("📋 Оставить контакт",     callback_data="flow_contact")],
        [InlineKeyboardButton("💬 Написать менеджеру",   url=MANAGER_URL)],
    ])
    await message.reply_text(
        "📦 <b>Что войдёт в расчёт:</b>\n\n"
        "— цена авто в Корее\n"
        "— доставка до Владивостока\n"
        "— таможня и утильсбор\n"
        "— <b>итог в РФ</b>\n\n"
        "Оставь контакт — отправим подборку.",
        parse_mode=ParseMode.HTML,
        reply_markup=keyboard,
    )


async def cmd_reset(update: Update, context: ContextTypes.DEFAULT_TYPE):
    clear_history(update.effective_chat.id)
    await update.message.reply_text(
        "🔄 Диалог сброшен. Начинаем заново!\n"
        "Напишите что ищете — например: <i>«Кроссовер до 4 миллионов, Тюмень»</i>",
        parse_mode=ParseMode.HTML,
    )


async def cmd_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "ℹ️ <b>Как работает агент:</b>\n\n"
        "Просто напишите что ищете — я задам уточняющие вопросы, "
        "найду варианты на <b>Encar.com</b> и посчитаю полную стоимость "
        "<b>под ключ</b> с доставкой до вашего города.\n\n"
        "Например:\n"
        "• <i>«Хочу кроссовер до 4 млн, Новосибирск»</i>\n"
        "• <i>«Что можно взять за 3 миллиона? Тюмень»</i>\n"
        "• <i>«Kia Sportage 2022, сколько выйдет в Екатеринбурге?»</i>\n\n"
        "<b>Команды клиента:</b>\n"
        "/start — начать заново\n"
        "/reset — сбросить диалог\n"
        "/watchlist — мой список отслеживания\n\n"
        "<b>Команды менеджера:</b>\n"
        "/leads — активные заявки\n"
        "/stats — статистика\n"
        "/export — выгрузить заявки в Excel\n"
        "/reply <i>chat_id текст</i> — ответить клиенту",
        parse_mode=ParseMode.HTML,
    )


async def cmd_watchlist(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    all_wl  = get_active_watchlist()
    mine    = [w for w in all_wl if w["chat_id"] == chat_id]

    if not mine:
        await update.message.reply_text(
            "📋 Ваш список отслеживания пуст.\n\n"
            "Напишите мне что ищете, и я поставлю авто на мониторинг — "
            "уведомлю как только появится подходящий вариант.",
            parse_mode=ParseMode.HTML,
        )
        return

    lines = ["📋 <b>Ваш список отслеживания:</b>\n"]
    for i, w in enumerate(mine, 1):
        lines.append(
            f"{i}. {w.get('maker', '') or '—'} {w.get('model', '') or ''} · "
            f"до {int(w['budget_rub']):,} ₽ · {w['city']}"
        )
    lines.append("\nПолучите уведомление как только появится подходящий вариант на Encar.")

    await update.message.reply_text(
        "\n".join(lines),
        parse_mode=ParseMode.HTML,
    )


async def cmd_leads(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = str(update.effective_chat.id)
    if MANAGER_CHAT_ID and chat_id != MANAGER_CHAT_ID:
        await update.message.reply_text("⛔ Команда доступна только менеджеру.")
        return

    leads_new  = get_leads(status="new")
    leads_work = get_leads(status="in_work")
    leads = leads_new + leads_work

    if not leads:
        await update.message.reply_text("📭 Активных заявок нет.")
        return

    await update.message.reply_text(
        f"📬 <b>Активные заявки: {len(leads)}</b> "
        f"(🆕 новых: {len(leads_new)} · 🔄 в работе: {len(leads_work)})",
        parse_mode=ParseMode.HTML,
    )

    for l in leads:
        status_icon = "🆕" if l["status"] == "new" else "🔄"
        text = (
            f"{status_icon} <b>#{l['id']} · {l['name']}</b>\n"
            f"📞 {l['phone']}\n"
            f"🚗 {l['car_choice']}\n"
            f"💰 {int(l['budget_rub'] or 0):,} ₽\n"
            f"🕐 {l['created_at'][:16]}"
        )
        keyboard = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("🔄 В работу",  callback_data=f"lead_work_{l['id']}"),
                InlineKeyboardButton("✅ Закрыть",   callback_data=f"lead_done_{l['id']}"),
            ],
            [
                InlineKeyboardButton("💬 Ответить клиенту", callback_data=f"lead_reply_{l['id']}_{l['chat_id']}"),
            ],
        ])
        await update.message.reply_text(text, parse_mode=ParseMode.HTML, reply_markup=keyboard)


async def cmd_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = str(update.effective_chat.id)
    if MANAGER_CHAT_ID and chat_id != MANAGER_CHAT_ID:
        await update.message.reply_text("⛔ Команда доступна только менеджеру.")
        return

    s = get_stats()
    conv = round(s["leads_total"] / max(s["total_clients"], 1) * 100, 1)

    lines = [
        "📊 <b>Статистика DriveVostoka</b>\n",
        f"👥 Клиентов всего: <b>{s['total_clients']}</b>",
        f"📨 Заявок всего: <b>{s['leads_total']}</b>  (конверсия <b>{conv}%</b>)",
        f"",
        f"🆕 Новых: <b>{s['leads_new']}</b>",
        f"🔄 В работе: <b>{s['leads_in_work']}</b>",
        f"✅ Закрытых: <b>{s['leads_done']}</b>",
        f"",
        f"📅 Сегодня: <b>{s['today']}</b> заявок",
        f"📅 За 7 дней: <b>{s['week']}</b> заявок",
    ]

    if s["top_models"]:
        lines.append("\n🏆 <b>Топ запросов:</b>")
        for i, (model, cnt) in enumerate(s["top_models"], 1):
            lines.append(f"  {i}. {model} — {cnt}")

    await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.HTML)


async def cmd_export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = str(update.effective_chat.id)
    if MANAGER_CHAT_ID and chat_id != MANAGER_CHAT_ID:
        await update.message.reply_text("⛔ Команда доступна только менеджеру.")
        return

    msg = await update.message.reply_text("⏳ Формирую Excel...")
    try:
        xlsx_bytes = _build_leads_xlsx()
        await context.bot.send_document(
            chat_id=update.effective_chat.id,
            document=BytesIO(xlsx_bytes),
            filename="drivevostoka_leads.xlsx",
            caption="📊 Все заявки DriveVostoka",
        )
    except Exception as e:
        logger.error(f"Export error: {e}")
        await msg.edit_text("❌ Ошибка при создании файла.")
        return
    await msg.delete()


def _build_leads_xlsx() -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("Установите openpyxl: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = ["ID", "Дата", "Имя", "Телефон", "Автомобиль", "Бюджет ₽", "Статус", "chat_id"]
    header_fill = PatternFill("solid", fgColor="1E6FBF")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    status_ru = {"new": "Новая", "in_work": "В работе", "done": "Закрыта"}
    leads = get_all_leads()

    for row_idx, lead in enumerate(leads, 2):
        ws.cell(row=row_idx, column=1, value=lead["id"])
        ws.cell(row=row_idx, column=2, value=lead["created_at"][:16])
        ws.cell(row=row_idx, column=3, value=lead["name"])
        ws.cell(row=row_idx, column=4, value=lead["phone"])
        ws.cell(row=row_idx, column=5, value=lead["car_choice"])
        ws.cell(row=row_idx, column=6, value=int(lead["budget_rub"] or 0))
        ws.cell(row=row_idx, column=7, value=status_ru.get(lead["status"], lead["status"]))
        ws.cell(row=row_idx, column=8, value=lead["chat_id"])

    col_widths = [6, 18, 20, 16, 40, 14, 12, 14]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def cmd_reply(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = str(update.effective_chat.id)
    if MANAGER_CHAT_ID and chat_id != MANAGER_CHAT_ID:
        await update.message.reply_text("⛔ Команда доступна только менеджеру.")
        return

    args = context.args
    if not args or len(args) < 2:
        await update.message.reply_text(
            "⚠️ Формат: <code>/reply 123456789 Текст сообщения</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    try:
        target_chat_id = int(args[0])
    except ValueError:
        await update.message.reply_text("⚠️ Неверный chat_id")
        return

    text = " ".join(args[1:])
    try:
        await context.bot.send_message(
            chat_id=target_chat_id,
            text=f"💬 <b>Сообщение от менеджера DriveVostoka:</b>\n\n{text}",
            parse_mode=ParseMode.HTML,
        )
        await update.message.reply_text(f"✅ Сообщение отправлено клиенту {target_chat_id}")
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка: {e}")


async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query   = update.callback_query
    chat_id = query.message.chat_id
    data    = query.data
    await query.answer()

    if data.startswith("lead_work_") or data.startswith("lead_done_"):
        parts     = data.split("_")
        new_status = "in_work" if parts[1] == "work" else "done"
        lead_id    = int(parts[2])
        update_lead_status(lead_id, new_status)
        status_text = "🔄 В работе" if new_status == "in_work" else "✅ Закрыта"
        await query.edit_message_reply_markup(reply_markup=None)
        await query.message.reply_text(
            f"Заявка <b>#{lead_id}</b> → {status_text}",
            parse_mode=ParseMode.HTML,
        )
        return

    if data.startswith("lead_reply_"):
        parts = data.split("_")
        lead_id       = parts[2]
        client_chat   = parts[3]
        await query.message.reply_text(
            f"✏️ Напишите ответ командой:\n"
            f"<code>/reply {client_chat} Ваш текст здесь</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    if data == "flow_start":
        await _show_budget_step(query.message)
        return

    if data == "flow_delivery":
        await query.message.reply_text(
            "🚢 <b>Как проходит доставка из Кореи:</b>\n\n"
            "1️⃣ Выбираем авто на аукционе / у дилера в Корее\n"
            "2️⃣ Проверка авто по VIN — история, ДТП, пробег\n"
            "3️⃣ Выкуп и отправка морем до Владивостока\n"
            "4️⃣ Таможенное оформление во Владивостоке\n"
            "5️⃣ Выдача авто с документами под ключ\n\n"
            "⏱ <b>Срок:</b> 30–45 дней от момента выкупа\n"
            "📍 <b>Точка выдачи:</b> Владивосток\n\n"
            "Всё включено — никаких скрытых платежей.",
            parse_mode=ParseMode.HTML,
        )
        await _show_budget_step(query.message)
        return

    if data in BUDGET_MAP:
        budget_max, budget_label = BUDGET_MAP[data]
        context.user_data["budget_max"]   = budget_max
        context.user_data["budget_label"] = budget_label
        save_profile(chat_id, budget_rub=budget_max, budget_label=budget_label)
        await _show_type_step(query.message)
        return

    if data in TYPE_MAP:
        type_label, body_type, is_electric = TYPE_MAP[data]
        context.user_data["type_label"]  = type_label
        context.user_data["body_type"]   = body_type
        context.user_data["is_electric"] = is_electric
        save_profile(chat_id, body_type=body_type,
                     engine_type="электро" if is_electric else "бензин")
        await _show_goal_step(query.message)
        return

    if data in GOAL_MAP:
        goal = GOAL_MAP[data]
        context.user_data["goal"] = goal
        save_profile(chat_id, goal=goal)
        budget_label = context.user_data.get("budget_label", "выбранный")
        type_label   = context.user_data.get("type_label",   "авто")
        await _show_warmup(query.message, budget_label, type_label, goal)
        return

    if data == "flow_getlist":
        budget_max   = context.user_data.get("budget_max",   3_000_000)
        budget_label = context.user_data.get("budget_label", "указанный")
        body_type    = context.user_data.get("body_type",    "")
        is_electric  = context.user_data.get("is_electric",  False)
        goal         = context.user_data.get("goal",         "для себя")

        goal_hint = {
            "для себя":            "Клиент ищет авто для себя — важны комплектация и надёжность.",
            "на перепродажу":      "Клиент ищет на перепродажу — важны ликвидность и маржа.",
            "сравнить с рынком РФ":"Клиент хочет сравнить с рынком РФ — покажи разницу в цене.",
        }.get(goal, "")

        search_msg = (
            f"Подбери 3 лучших варианта под параметры:\n"
            f"— Бюджет: {budget_label} (до {budget_max:,} ₽ до Владивостока)\n"
            f"— Тип: {body_type or 'любой'}\n"
            f"— Электромобиль: {'да' if is_electric else 'нет'}\n"
            f"— Цель: {goal}\n"
            f"{goal_hint}\n"
            f"Выполни поиск и покажи сравнительную таблицу с ценой до Владивостока."
        )

        msg = await query.message.reply_text("⏳ Ищу варианты под твои параметры...")
        response = await _process(chat_id, search_msg)
        text, photo_urls = _extract_photos(response)

        if photo_urls:
            await _send_photos(query.message, photo_urls)

        await msg.edit_text(text, parse_mode=ParseMode.HTML, disable_web_page_preview=True)
        await _show_lead_block(query.message)
        return

    if data == "flow_contact":
        contact_msg = (
            "Клиент хочет оставить контакт для получения подборки. "
            "Спроси имя и номер телефона, затем зарегистрируй заявку через register_lead."
        )
        msg = await query.message.reply_text("⏳")
        response = await _process(chat_id, contact_msg)
        text, _ = _extract_photos(response)
        await msg.edit_text(text, parse_mode=ParseMode.HTML, disable_web_page_preview=True)
        return

    legacy = {
        "find_car": "Хочу подобрать автомобиль из Кореи.",
        "calc":     "Хочу рассчитать стоимость ввоза конкретного автомобиля.",
    }
    user_text = legacy.get(data)
    if not user_text:
        return

    msg = await query.message.reply_text("⏳ Ищу варианты...", parse_mode=ParseMode.HTML)
    response = await _process(chat_id, user_text)
    text, photo_urls = _extract_photos(response)

    if photo_urls:
        await _send_photos(query.message, photo_urls)

    await msg.edit_text(text, parse_mode=ParseMode.HTML, disable_web_page_preview=True)


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id   = update.effective_chat.id
    user_text = update.message.text

    upsert_client(
        chat_id=chat_id,
        username=update.effective_user.username or "",
        first_name=update.effective_user.first_name or "",
    )

    typing = await update.message.reply_text("⏳ Ищу варианты...")

    response = await _process(chat_id, user_text)
    text, photo_urls = _extract_photos(response)

    try:
        await typing.delete()
    except Exception:
        pass

    if photo_urls:
        await _send_photos(update.message, photo_urls)

    for chunk in _split(text, 4000):
        await update.message.reply_text(
            chunk,
            parse_mode=ParseMode.HTML,
            disable_web_page_preview=True,
        )

    if "Заявка принята" in text and MANAGER_CHAT_ID:
        try:
            await context.bot.send_message(
                chat_id=int(MANAGER_CHAT_ID),
                text=f"🔔 <b>Новая заявка!</b>\nchat_id: {chat_id}\n\n{text}",
                parse_mode=ParseMode.HTML,
            )
        except Exception as e:
            logger.error(f"Ошибка отправки менеджеру: {e}")


async def _send_photos(message, photo_urls: list[str]) -> None:
    try:
        photos_bytes = await prepare_photos(photo_urls, limit=2)
        if not photos_bytes:
            return
        from io import BytesIO
        media = [InputMediaPhoto(media=BytesIO(b)) for b in photos_bytes]
        await message.reply_media_group(media=media)
    except Exception as e:
        logger.warning(f"[Photos] Ошибка отправки медиагруппы: {e}")


async def _process(chat_id: int, text: str) -> str:
    try:
        return await run_agent(chat_id, text, claude)
    except Exception as e:
        logger.error(f"Agent error: {e}")
        return "😔 Что-то пошло не так. Попробуйте ещё раз или напишите менеджеру: @DriveVostoka"


_PHOTOS_RE = re.compile(r'^\[PHOTOS:([^\]]+)\]\s*', re.MULTILINE)


def _extract_photos(response: str) -> tuple[str, list[str]]:
    match = _PHOTOS_RE.search(response)
    if not match:
        return response, []
    urls = [u.strip() for u in match.group(1).split("||") if u.strip()]
    clean = _PHOTOS_RE.sub("", response).strip()
    return clean, urls


def _split(text: str, size: int) -> list[str]:
    return [text[i:i+size] for i in range(0, len(text), size)]


async def post_init(application: Application):
    asyncio.create_task(run_monitor(application.bot))
    logger.info("✅ Автомониторинг Encar запущен как фоновая задача")


def main():
    init_db()
    migrate_db()
    logger.info("✅ База данных инициализирована и обновлена")

    app = (
        Application.builder()
        .token(TG_TOKEN)
        .post_init(post_init)
        .build()
    )

    app.add_handler(CommandHandler("start",     cmd_start))
    app.add_handler(CommandHandler("reset",     cmd_reset))
    app.add_handler(CommandHandler("help",      cmd_help))
    app.add_handler(CommandHandler("watchlist", cmd_watchlist))
    app.add_handler(CommandHandler("leads",     cmd_leads))
    app.add_handler(CommandHandler("stats",     cmd_stats))
    app.add_handler(CommandHandler("export",    cmd_export))
    app.add_handler(CommandHandler("reply",     cmd_reply))
    app.add_handler(CallbackQueryHandler(handle_callback))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    logger.info("🚀 DriveVostoka Agent запущен")
    app.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    main()
